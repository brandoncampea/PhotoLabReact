import openpyxl
import csv

wb = openpyxl.load_workbook('2024 WHCC List Pricing.xlsx')
sheet_names = wb.sheetnames

with open('whcc_all_products_full.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Sheet', 'Product Code', 'Product Name/Size', 'Price'])
    for sheet in sheet_names:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2):
            code = row[0].value
            name = row[1].value
            price = row[2].value
            if code and name and price:
                writer.writerow([sheet, code, name.strip(), price])
